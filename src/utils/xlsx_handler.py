import calendar
import datetime
import openpyxl
try:
    from xlsx_formats import *
except ModuleNotFoundError:
    from utils.xlsx_formats import *
    
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import traceback
import os

def remake_xlsx_file(xlsx_path: str = "data/purchases.xlsx") -> None:
    """Remake the workbook at `xlsx_path` for the app.

    Actions:
    - Load workbook and create a timestamped backup in the same folder.
    - Ensure the first worksheet has headers and columns initialized via xlsx_init_column.
    - Sort existing data rows by the Date column (oldest -> newest).
    - Recreate category data validation and conditional formatting.
    - Apply row-level formatting via xlsx_format_rows and save the workbook.

    The function is defensive but will raise exceptions for missing files so callers
    can handle errors as needed.
    """
    
    try:
        # Basic validation and debugging info
        abs_path = os.path.abspath(xlsx_path)
        
        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"Workbook not found: {xlsx_path} (absolute: {abs_path})")

        # Test if openpyxl works at all in this environment
        try:
            test_wb = openpyxl.Workbook()
            test_ws = test_wb.active
            test_ws['A1'] = 'test'
        except Exception as test_e:
            raise RuntimeError(f"openpyxl basic functionality test failed: {type(test_e).__name__}: {test_e}\nTraceback: {traceback.format_exc()}")
        
        # Load the actual workbook
        wb = openpyxl.load_workbook(xlsx_path)
        
        if wb is None:
            raise ValueError(f"openpyxl.load_workbook returned None for file: {abs_path}")
            
    except Exception as e:
        error_context = f"""
            Environment Info:
            - Path: {xlsx_path}
            - Absolute path: {os.path.abspath(xlsx_path) if os.path else 'N/A'}
            - File exists: {os.path.exists(xlsx_path) if os.path else 'N/A'}
            - Current working directory: {os.getcwd() if os else 'N/A'}
            - openpyxl version: {getattr(openpyxl, '__version__', 'unknown')}
            - Error type: {type(e).__name__}
            - Error message: {str(e)}

            Full traceback:
            {traceback.format_exc()}"""
        raise RuntimeError(f"Failed to load workbook: {error_context}") from e

    # Backup
    try:
        now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        folder = os.path.dirname(xlsx_path) or "."
        base = os.path.basename(xlsx_path)
        backup_name = f"{os.path.splitext(base)[0]}_backup_{now}.xlsx"
        backup_path = os.path.join(folder, backup_name)
        
        wb.save(backup_path)
    except Exception as e:
        raise RuntimeError(f"Failed to create backup at '{backup_path}': {type(e).__name__}: {e}\nTraceback: {traceback.format_exc()}") from e

    try:
        ws = wb.worksheets[0]
    except Exception as e:
        raise RuntimeError(f"Failed to get worksheet: {type(e).__name__}: {e}\nTraceback: {traceback.format_exc()}") from e

    # Determine header row (assume row 1) and map headers to columns
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        cell = ws[f"{letter}1"]
        if cell.value:
            headers[str(cell.value).strip()] = letter

    # If expected headers are absent, we'll initialize columns A-H using defaults similar to the script
    default_cols = {
        "Item": ["A", 15],
        "Category": ["B", 15],
        "Cost": ["C", 10],
        "Date": ["D", 12],
        "Notes": ["E", 15],
        "Month": ["F", 12],
        "MonthNum": ["G", 5],
        "Year": ["H", 10]
    }

    # Ensure these headers exist in the sheet; xlsx_init_column will set header cell values
    cols = {}
    for key, (letter, width) in default_cols.items():
        cols[key] = [letter, width]
        xlsx_init_column(ws, letter, key, width)

    # Collect data rows (from row 2 to last non-empty row) and sort by Date
    data_rows = []
    max_row = ws.max_row
    # read values for all columns currently used (we'll keep the full row as a list)
    for r in range(2, max_row + 1):
        # Determine if the row is empty by checking the Item column (A)
        item_cell = ws[f"A{r}"]
        if item_cell.value is None:
            continue
        row_values = []
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            row_values.append(ws[f"{letter}{r}"].value)
        data_rows.append(row_values)

    # Attempt to identify which column index corresponds to Date (by header name)
    date_col_letter = cols.get('Date')[0]
    date_col_index = None
    try:
        # convert letter to 0-based index
        date_col_index = openpyxl.utils.column_index_from_string(date_col_letter) - 1
    except Exception:
        date_col_index = None

    # Parse and sort rows by date if possible; otherwise keep original order
    def _parse_date(val):
        try:
            if isinstance(val, datetime.datetime) or isinstance(val, datetime.date):
                return datetime.datetime(val.year, val.month, val.day)
            return datetime.datetime.fromisoformat(str(val))
        except Exception:
            try:
                # fallback to pandas parser if available
                import dateutil.parser as _dp
                return _dp.parse(str(val))
            except Exception:
                return None

    if date_col_index is not None:
        sortable = []
        nonsortable = []
        for row in data_rows:
            try:
                d = row[date_col_index]
            except Exception:
                d = None
            parsed = _parse_date(d)
            if parsed is not None:
                # replace the date cell value with the parsed datetime so Excel
                # interprets it as a date rather than a string when we write it back
                row_copy = list(row)
                row_copy[date_col_index] = parsed
                sortable.append((parsed, row_copy))
            else:
                nonsortable.append(row)

        # sort by parsed date (oldest -> newest)
        sortable.sort(key=lambda x: x[0])
        sorted_rows = [r for _, r in sortable] + nonsortable
    else:
        sorted_rows = data_rows

    # Clear existing data rows (from row 2 onward) to rewrite sorted rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            ws[f"{letter}{r}"] = None

    # Write back sorted rows starting at row 2
    write_row = 2
    for row in sorted_rows:
        for idx, val in enumerate(row, start=1):
            letter = get_column_letter(idx)
            ws[f"{letter}{write_row}"] = val
        write_row += 1
        
    # Reset autofilter across A:H
    try:
        ws.auto_filter.ref = "A1:H1"
    except Exception:
        pass

    # Recreate category DV and CF using column B
    xlsx_create_category_dv(ws, "B")
    xlsx_create_category_cf(ws, "B")

    # Apply row formatting based on our cols mapping
    xlsx_format_rows(ws, cols)

    # Save workbook with error handling
    try:
        if wb is None:
            raise ValueError("Workbook is None before save operation")
        wb.save(xlsx_path)
    except Exception as e:
        raise RuntimeError(f"Failed to save workbook to '{xlsx_path}': {type(e).__name__}: {e}\nTraceback: {traceback.format_exc()}") from e

def xlsx_init_column(ws: Worksheet, col_letter: str, text: str, width: float):
    # Create Header Object
    header_cell: Cell = ws[f"{col_letter}{1}"]
    header_cell.value = text
    header_cell.font = FONT_BOLD
    header_cell.alignment = ALIGN_CENTER
    header_cell.border = BORDER_ALL
    
    # Set width
    col_obj = ws.column_dimensions[col_letter]
    col_obj.width = width
    
    
def xlsx_format_rows(ws: Worksheet, cols: dict):
    item_col = cols.get("Item")[0]
    date_col = cols.get("Date")[0]
    category_col = cols.get("Category")[0]
    
    for row in range(2, ws.max_row + 150):
        # We're tracking if there is an entry in this part so we can continue adding formatting if there is no data
        item_cell = ws[f"{item_col}{row}"]
        is_empty = item_cell.value is None
        
        # Sets the fill of the cell based on the month
        if not is_empty:
            date_cell: Cell = ws[f"{date_col}{row}"]
            date_value: datetime = date_cell.value
            
            # Handle case where item exists but date is None/empty
            if date_value is not None:
                try:
                    month = date_value.month
                    year = date_value.year
                    fill = FILL_WHITE if month % 2 == 0 else FILL_GREY
                except AttributeError:
                    # date_value is not a datetime object
                    fill = FILL_WHITE
            else:
                # date_value is None
                fill = FILL_WHITE
        else:
            fill = FILL_WHITE
        
        # If the item is a subscription, change the fill to yellow
        category_cell = ws[f"{category_col}{row}"]
        category_value = category_cell.value
        if is_empty:
            pass
        elif category_value == "Digital Subscriptions":
            fill = FILL_SUBSCRIPTION
        
        # Update values in each individual cell
        for col, (col_letter, width) in cols.items():
            cell: Cell = ws[f"{col_letter}{row}"]
            cell.border = BORDER_LEFT_RIGHT
            # Default variables
            cell.fill = fill
            cell.alignment = ALIGN_CENTER 
            
            if col == "Item":
                cell.alignment = ALIGN_LEFT
            elif col == "Notes":
                cell.alignment = ALIGN_LEFT
            elif col == "Category":
                pass
            elif col == "Cost":
                cell.number_format = FORMAT_MONEY
            elif col == "Date":
                cell.number_format = FORMAT_DATE
            elif not is_empty:
                if col == "Month":
                    cell.value = calendar.month_name[month]
                elif col == "MonthNum":
                    cell.value = month
                elif col == "Year":
                    cell.value = year
                

def xlsx_create_category_dv(ws: Worksheet, category_col: str = "B"):
    # Create a string for the formula (must be double-quoted and comma-separated)
    category_list = '"' + ",".join(CATEGORIES) + '"'

    # Create a DataValidation object
    dv = DataValidation(type="list", formula1=category_list, allow_blank=True)

    # Add the DataValidation to the worksheet
    ws.add_data_validation(dv)

    # Apply it to B2:B<max_row>
    dv.add(f"{category_col}2:{category_col}{(int)(ws.max_row*2)}")
    

def xlsx_create_category_cf(ws: Worksheet, category_col: str = "B"):
    # Apply conditional formatting based on FILL_MAP
    for category_name, pattern_fill in FILL_MAP.items():
        color = pattern_fill.fgColor.rgb  # Get RGB value

        # Create a formula rule
        formula = f'$B2="{category_name}"'  # Important: Lock column B, but row can move
        rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))

        # Add the rule to the worksheet
        ws.conditional_formatting.add(f"{category_col}2:{category_col}{ws.max_row*2}", rule)


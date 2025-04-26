from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

CATEGORIES = [
    "Food & Beverages",
    "Books & Literature",
    "Gaming",
    "Digital Subscriptions",
    "Movies & Media",
    "Music & Audio",
    "Electronics & Accessories",
    "Clothing & Apparel",
    "Health & Personal Care",
    "Collectibles",
    "Miscellaneous"
]

# Create a white fill style
FILL_WHITE = PatternFill(fgColor="FFFFFF", fill_type="solid")
FILL_GREY = PatternFill(fgColor="d9d9d9", fill_type="solid")
FILL_SUBSCRIPTION = PatternFill(fgColor="fbffa9", fill_type="solid")

FILL_MAP = {
    "Electronics & Accessories": PatternFill(fgColor="729fcf", fill_type="solid"),
    "Movies & Media": PatternFill(fgColor="a05eff", fill_type="solid"),
    "Books & Literature": PatternFill(fgColor="e9a9ff", fill_type="solid"),
    "Gaming": PatternFill(fgColor="bbe33d", fill_type="solid"),
    "Collectibles": PatternFill(fgColor="ffc85d", fill_type="solid"),
    "Music & Audio": PatternFill(fgColor="ffa9f2", fill_type="solid"),
    "Health & Personal Care": PatternFill(fgColor="a9ffc4", fill_type="solid"),
    "Digital Subscriptions": PatternFill(fgColor="fbffa9", fill_type="solid")
}

# Create thin border style
SIDE_THIN = Side(style="thin")
# Define border with left and right
BORDER_LEFT_RIGHT = Border(left=SIDE_THIN, right=SIDE_THIN)
BORDER_ALL = Border(left=SIDE_THIN, right=SIDE_THIN, top=SIDE_THIN, bottom=SIDE_THIN)

# Define center alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

# Define a bold font
FONT_BOLD = Font(bold=True)

# Your custom number formats
FORMAT_MONEY = '"$"* #,##0.00'
FORMAT_DATE = 'dd/mm/yyyy'


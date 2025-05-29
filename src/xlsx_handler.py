import calendar
import datetime
from tqdm import tqdm
try:
    from xlsx_formats import *
except ModuleNotFoundError:
    from src.xlsx_formats import *
    
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import FormulaRule

def xlsx_init_column(ws: Worksheet, col_letter: str, text: str, width: float):
    # Create Header Object
    header_cell: Cell = ws[f"{col_letter}{1}"]
    header_cell.value = text
    header_cell.font = FONT_BOLD
    header_cell.alignment = ALIGN_CENTER
    header_cell.border = BORDER_ALL
    
    # Set width
    col_obj = ws.column_dimensions[col_letter]
    col_obj.width = width
    
    
def xlsx_format_rows(ws: Worksheet, cols: dict):
    item_col = cols.get("Item")[0]
    date_col = cols.get("Date")[0]
    category_col = cols.get("Category")[0]
    
    for row in tqdm(range(2, ws.max_row + 150)):
        # We're tracking if there is an entry in this part so we can continue adding formatting if there is no data
        item_cell = ws[f"{item_col}{row}"]
        is_empty = item_cell.value is None
        
        # Sets the fill of the cell based on the month
        if not is_empty:
            date_cell: Cell = ws[f"{date_col}{row}"]
            date_value: datetime = date_cell.value
            month = date_value.month
            year = date_value.year
            fill = FILL_WHITE if month % 2 == 0 else FILL_GREY
            
        else:
            fill = FILL_WHITE
        
        # If the item is a subscription, change the fill to yellow
        category_cell = ws[f"{category_col}{row}"]
        category_value = category_cell.value
        if is_empty:
            pass
        elif category_value == "Digital Subscriptions":
            fill = FILL_SUBSCRIPTION
        
        # Update values in each individual cell
        for col, (col_letter, width) in cols.items():
            cell: Cell = ws[f"{col_letter}{row}"]
            cell.border = BORDER_LEFT_RIGHT
            # Default variables
            cell.fill = fill
            cell.alignment = ALIGN_CENTER 
            
            if col == "Item":
                cell.alignment = ALIGN_LEFT
            elif col == "Notes":
                cell.alignment = ALIGN_LEFT
            elif col == "Category":
                pass
            elif col == "Cost":
                cell.number_format = FORMAT_MONEY
            elif col == "Date":
                cell.number_format = FORMAT_DATE
            elif not is_empty:
                if col == "Month":
                    cell.value = calendar.month_name[month]
                elif col == "MonthNum":
                    cell.value = month
                elif col == "Year":
                    cell.value = year
                

def xlsx_create_category_dv(ws: Worksheet, category_col: str = "B"):
    # Create a string for the formula (must be double-quoted and comma-separated)
    category_list = '"' + ",".join(CATEGORIES) + '"'

    # Create a DataValidation object
    dv = DataValidation(type="list", formula1=category_list, allow_blank=True)

    # Add the DataValidation to the worksheet
    ws.add_data_validation(dv)

    # Apply it to B2:B<max_row>
    dv.add(f"{category_col}2:{category_col}{(int)(ws.max_row*2)}")
    

def xlsx_create_category_cf(ws: Worksheet, category_col: str = "B"):
    # Apply conditional formatting based on FILL_MAP
    for category_name, pattern_fill in FILL_MAP.items():
        color = pattern_fill.fgColor.rgb  # Get RGB value

        # Create a formula rule
        formula = f'$B2="{category_name}"'  # Important: Lock column B, but row can move
        rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))

        # Add the rule to the worksheet
        ws.conditional_formatting.add(f"{category_col}2:{category_col}{ws.max_row*2}", rule)

